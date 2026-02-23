import io
from openpyxl import load_workbook
from django.urls import reverse
from .base import BaseTestCase
from .utils import create_superuser, seed_language_lesson_cards

class AdminExportTests(BaseTestCase):
    def setUp(self):
        self.admin = create_superuser()
        seed_language_lesson_cards(lang_code="en", cards=1)

    def test_export_requires_admin(self):
        url = reverse("admin_report_export")
        r = self.client.get(url)
        self.assertIn(r.status_code, (302, 403))

    def test_export_returns_valid_xlsx(self):
        self.client.login(username="admin", password="Pass12345!")
        url = reverse("admin_report_export")

        post_data = {
            "models": ["Language"],
            "fields__Language": ["id", "name", "code"],
        }

        r = self.client.post(url, data=post_data)
        self.assertEqual(r.status_code, 200)
        self.assertTrue(r.content.startswith(b"PK\x03\x04"))

        wb = load_workbook(filename=io.BytesIO(r.content))
        self.assertGreaterEqual(len(wb.sheetnames), 1)